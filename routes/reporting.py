# -*- coding: utf-8 -*-
"""
Reporting routes - export dat z výpočtů koncových cen
"""

from flask import Blueprint, render_template, request, session, redirect, url_for, flash, make_response
from models import db, Stredisko, ObdobiFakturace, VypocetOM, OdberneMisto
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

reporting_bp = Blueprint('reporting', __name__, url_prefix='/reporting')

@reporting_bp.route('/')
def index():
    """Zobrazí stránku reportingu s možností výběru období, středisek a metrik"""
    if not session.get("user_id"):
        return redirect("/login")

    user_id = session["user_id"]

    # Načti všechna střediska uživatele
    strediska = Stredisko.query.filter_by(user_id=user_id).order_by(Stredisko.nazev_strediska).all()

    if not strediska:
        flash("❌ Nemáte vytvořena žádná střediska.", "warning")
        return redirect("/strediska")

    # Načti všechna období (pro výběr)
    obdobi_list = []
    for stredisko in strediska:
        obdobi = ObdobiFakturace.query.filter_by(stredisko_id=stredisko.id)\
            .order_by(ObdobiFakturace.rok.asc(), ObdobiFakturace.mesic.asc())\
            .all()
        for obj in obdobi:
            key = f"{obj.rok}-{obj.mesic:02d}"
            if key not in [o['key'] for o in obdobi_list]:
                obdobi_list.append({
                    'key': key,
                    'rok': obj.rok,
                    'mesic': obj.mesic,
                    'label': f"{obj.mesic:02d}/{obj.rok}"
                })

    # Seřaď období podle roku a měsíce (vzestupně)
    obdobi_list.sort(key=lambda x: (x['rok'], x['mesic']), reverse=False)

    # Definuj dostupné metriky z tabulky vypocty_om
    metriky = [
        {'name': 'spotreba_om', 'label': 'Spotřeba (kWh)'},
        {'name': 'platba_za_jistic', 'label': 'DIST. Platba za jistič (Kč)'},
        {'name': 'platba_za_distribuci_vt', 'label': 'DIST. Platba za distribuci VT (Kč)'},
        {'name': 'platba_za_distribuci_nt', 'label': 'DIST. Platba za distribuci NT (Kč)'},
        {'name': 'systemove_sluzby', 'label': 'DIST. Systémové služby (Kč)'},
        {'name': 'poze_dle_jistice', 'label': 'DIST. POZE dle jističe (Kč)'},
        {'name': 'poze_dle_spotreby', 'label': 'DIST. POZE dle spotřeby (Kč)'},
        {'name': 'nesitova_infrastruktura', 'label': 'DIST. Nesíťová infrastruktura (Kč)'},
        {'name': 'dan_z_elektriny', 'label': 'DOD. Daň z elektřiny (Kč)'},
        {'name': 'platba_za_elektrinu_vt', 'label': 'DOD. Silovina VT (Kč)'},
        {'name': 'platba_za_elektrinu_nt', 'label': 'DOD. Silovina NT (Kč)'},
        {'name': 'mesicni_plat', 'label': 'DOD. Měsíční plat (Kč)'},
        {'name': 'zaklad_bez_dph', 'label': 'Cena bez DPH (Kč)'},
        {'name': 'celkem_vc_dph', 'label': 'Cena vč. DPH (Kč)'},
    ]

    return render_template('reporting/reporting.html',
                         strediska=strediska,
                         obdobi_list=obdobi_list,
                         metriky=metriky)


@reporting_bp.route('/export', methods=['POST'])
def export():
    """Exportuje data do XLSX souboru podle výběru uživatele"""
    if not session.get("user_id"):
        return redirect("/login")

    user_id = session["user_id"]

    # Získej data z formuláře
    obdobi_str = request.form.get('obdobi')  # formát: "2025-09"
    vybrana_strediska_ids = request.form.getlist('strediska[]')  # seznam ID středisek
    vybrane_metriky = request.form.getlist('metriky[]')  # seznam názvů metrik

    # Validace
    if not obdobi_str:
        flash("❌ Vyberte období pro export.", "danger")
        return redirect(url_for('reporting.index'))

    if not vybrana_strediska_ids:
        flash("❌ Vyberte alespoň jedno středisko pro export.", "danger")
        return redirect(url_for('reporting.index'))

    if not vybrane_metriky:
        flash("❌ Vyberte alespoň jednu metriku pro export.", "danger")
        return redirect(url_for('reporting.index'))

    # Parsuj období
    try:
        rok, mesic = obdobi_str.split('-')
        rok = int(rok)
        mesic = int(mesic)
    except:
        flash("❌ Neplatný formát období.", "danger")
        return redirect(url_for('reporting.index'))

    # Převeď ID středisek na int
    try:
        strediska_ids = [int(sid) for sid in vybrana_strediska_ids]
    except:
        flash("❌ Neplatná ID středisek.", "danger")
        return redirect(url_for('reporting.index'))

    # Ověř že střediska patří uživateli
    strediska = Stredisko.query.filter(
        Stredisko.id.in_(strediska_ids),
        Stredisko.user_id == user_id
    ).all()

    if len(strediska) != len(strediska_ids):
        flash("❌ Některá střediska neexistují nebo k nim nemáte přístup.", "danger")
        return redirect(url_for('reporting.index'))

    # Najdi období pro každé středisko
    obdobi_ids = []
    for stredisko in strediska:
        obdobi = ObdobiFakturace.query.filter_by(
            stredisko_id=stredisko.id,
            rok=rok,
            mesic=mesic
        ).first()
        if obdobi:
            obdobi_ids.append(obdobi.id)

    if not obdobi_ids:
        flash(f"❌ Pro vybraná střediska neexistuje období {mesic:02d}/{rok}.", "danger")
        return redirect(url_for('reporting.index'))

    # Načti všechny výpočty pro vybraná období
    vypocty = db.session.query(VypocetOM, OdberneMisto)\
        .join(OdberneMisto, VypocetOM.odberne_misto_id == OdberneMisto.id)\
        .filter(VypocetOM.obdobi_id.in_(obdobi_ids))\
        .order_by(OdberneMisto.cislo_om)\
        .all()

    if not vypocty:
        flash(f"❌ Pro vybrané období a střediska neexistují žádné výpočty koncových cen.", "warning")
        return redirect(url_for('reporting.index'))

    # Vytvoř Excel soubor
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {mesic:02d}-{rok}"

    # Stylování
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Vytvoř záhlaví
    headers = ['Odběrné místo ID']

    # Mapování názvů metrik na lidsky čitelné názvy
    metriky_labels = {
        'spotreba_om': 'Spotřeba (kWh)',
        'platba_za_jistic': 'DIST. Platba za jistič (Kč)',
        'platba_za_distribuci_vt': 'DIST. Platba za distribuci VT (Kč)',
        'platba_za_distribuci_nt': 'DIST. Platba za distribuci NT (Kč)',
        'systemove_sluzby': 'DIST. Systémové služby (Kč)',
        'poze_dle_jistice': 'DIST. POZE dle jističe (Kč)',
        'poze_dle_spotreby': 'DIST. POZE dle spotřeby (Kč)',
        'nesitova_infrastruktura': 'DIST. Nesíťová infrastruktura (Kč)',
        'dan_z_elektriny': 'DOD. Daň z elektřiny (Kč)',
        'platba_za_elektrinu_vt': 'DOD. Silovina VT (Kč)',
        'platba_za_elektrinu_nt': 'DOD. Silovina NT (Kč)',
        'mesicni_plat': 'DOD. Měsíční plat (Kč)',
        'zaklad_bez_dph': 'Cena bez DPH (Kč)',
        'celkem_vc_dph': 'Cena vč. DPH (Kč)',
    }

    for metrika in vybrane_metriky:
        headers.append(metriky_labels.get(metrika, metrika))

    # Zapiš záhlaví
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Zapiš data
    row_num = 2
    for vypocet, om in vypocty:
        # První sloupec - ID odběrného místa
        ws.cell(row=row_num, column=1, value=om.cislo_om)

        # Další sloupce - vybrané metriky
        for col_num, metrika in enumerate(vybrane_metriky, 2):
            hodnota = getattr(vypocet, metrika, None)

            # Defaultní hodnota pokud není vyplněno
            if hodnota is None:
                hodnota = 0

            # Zapiš hodnotu jako číslo
            cell = ws.cell(row=row_num, column=col_num, value=float(hodnota))

            # Nastav formátování podle typu metriky
            if metrika == 'spotreba_om':
                # Spotřeba v kWh - custom formát s mezerami jako oddělovač tisíců
                cell.number_format = '### ###0" kWh"'
            else:
                # Všechny ostatní hodnoty v Kč - currency formát
                cell.number_format = '#,##0.00" Kč"'

        row_num += 1

    # Automatické šířky sloupců
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Vytvoř response s Excel souborem
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Název souboru
    filename = f"report_{mesic:02d}_{rok}.xlsx"

    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response
